"""
Step 9 – Output & Reporting
Generates PDF (ReportLab) and Excel (openpyxl) from real computed survey data.
No hardcoded fallback values — missing fields are shown as 'Not computed'.
"""
import os
from datetime import datetime, timezone

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

NA = "Not computed"

def _v(val, fmt=None, suffix=""):
    if val is None:
        return NA
    try:
        return (fmt % val if fmt else str(val)) + suffix
    except Exception:
        return NA

def generate_pdf(survey, site) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)

    fname = f"survey_{survey.id}_{int(datetime.now(timezone.utc).timestamp())}.pdf"
    path  = os.path.join(REPORTS_DIR, fname)

    DARK  = colors.HexColor("#0f1421")
    AMBER = colors.HexColor("#f59e0b")
    GRAY  = colors.HexColor("#64748b")
    LIGHT = colors.HexColor("#f8fafc")

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    brand   = ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=10, textColor=AMBER)
    title_s = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=22, textColor=DARK, spaceAfter=4)
    sub_s   = ParagraphStyle("sub",   fontName="Helvetica",      fontSize=11, textColor=GRAY, spaceAfter=16)
    h2_s    = ParagraphStyle("h2",    fontName="Helvetica-Bold", fontSize=13, textColor=DARK, spaceBefore=14, spaceAfter=6)
    foot_s  = ParagraphStyle("foot",  fontName="Helvetica",      fontSize=8,  textColor=GRAY)

    ts = TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("FONTNAME",  (0,0), (0,-1),  "Helvetica-Bold"),
        ("FONTNAME",  (2,0), (2,-1),  "Helvetica-Bold"),
        ("GRID",      (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, LIGHT]),
        ("PADDING",   (0,0), (-1,-1), 5),
        ("BACKGROUND",(0,0), (0,-1),  LIGHT),
        ("BACKGROUND",(2,0), (2,-1),  LIGHT),
    ])

    def meta_table(rows):
        t = Table(rows, colWidths=[42*mm, 52*mm, 42*mm, 44*mm])
        t.setStyle(ts)
        return t

    story = []
    story.append(Paragraph("MineVisionAI", brand))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Volume Survey Report", title_s))
    story.append(Paragraph(f"{site.name} · {site.location or ''}, {site.state or ''}", sub_s))
    story.append(HRFlowable(width="100%", thickness=2, color=AMBER))
    story.append(Spacer(1, 6*mm))

    now = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    story.append(meta_table([
        ["Report Date",     now,
         "Mine Type",       site.mine_type or NA],
        ["Drone Model",     survey.drone_model or NA,
         "Flying Height",   _v(survey.flying_height_m, "%.0f", " m")],
        ["Images Uploaded", _v(survey.image_count, "%d"),
         "Images Passed QC",_v(survey.images_passed, "%d")],
        ["Images Rejected", _v(survey.images_rejected, "%d"),
         "GCPs Used",       _v(survey.gcp_count, "%d")],
        ["DEM Resolution",  _v(survey.dem_resolution, "%.2f", " m/px"),
         "Coordinate System", survey.dem_crs or NA],
    ]))
    story.append(Spacer(1, 8*mm))

    # Volume results
    story.append(Paragraph("Volume Analysis Results", h2_s))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    story.append(Spacer(1, 3*mm))

    vol_rows = [
        ["Metric",            "Value",          "Unit",   "Method / Notes"],
        ["Stockpile Volume",  _v(survey.stockpile_volume, "%,.0f"), "m³",
         "TIN above reference plane"],
        ["Cut Volume",        _v(survey.cut_volume,       "%,.0f"), "m³", "Below reference plane"],
        ["Net Change",        _v(survey.net_change,       "%+,.0f") if survey.net_change else NA, "m³", "vs reference"],
        ["Reference Plane",   _v(survey.reference_elev,  "%.2f"),  "m ASL", "5th percentile of DEM"],
        ["Mine Area",         _v(site.area_km2,           "%.4f"),  "km²",  "From DEM valid pixels"],
        ["Max Pit Depth",     _v(site.max_depth_m,        "%.1f"),  "m",    "User input"],
        ["Elevation Min",     _v(site.elevation_min,      "%.1f"),  "m ASL","From DEM"],
        ["Elevation Max",     _v(site.elevation_max,      "%.1f"),  "m ASL","From DEM"],
    ]
    tv = Table(vol_rows, colWidths=[52*mm, 48*mm, 18*mm, 62*mm])
    tv.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), DARK),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME",     (0,1), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, LIGHT]),
        ("PADDING",      (0,0), (-1,-1), 6),
        ("ALIGN",        (1,0), (2,-1), "RIGHT"),
    ]))
    story.append(tv)
    story.append(Spacer(1, 8*mm))

    # Pipeline summary
    import json
    steps = json.loads(survey.pipeline_steps) if survey.pipeline_steps else []
    if steps:
        story.append(Paragraph("Processing Pipeline", h2_s))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
        story.append(Spacer(1, 3*mm))
        p_rows = [["Step", "Status", "Detail"]]
        for s in steps:
            icon = {"complete":"✓","failed":"✗","skipped":"—","running":"⟳"}.get(s["status"],"·")
            p_rows.append([s["name"], f"{icon} {s['status'].title()}", s.get("detail","")])
        tp = Table(p_rows, colWidths=[54*mm, 28*mm, 98*mm])
        tp.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), DARK),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("TEXTCOLOR",  (1,1), (1,-1), colors.HexColor("#16a34a")),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LIGHT]),
            ("PADDING",    (0,0), (-1,-1), 5),
        ]))
        story.append(tp)

    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=AMBER))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Volume calculations use TIN-based DEM differencing above the reference plane. "
        "Accuracy depends on GCP distribution, image overlap, and DEM resolution.",
        foot_s))

    doc.build(story)
    return path


def generate_excel(survey, site, db) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from database import VolumeHistory, UploadedFile
    import json

    fname = f"survey_{survey.id}_{int(datetime.now(timezone.utc).timestamp())}.xlsx"
    path  = os.path.join(REPORTS_DIR, fname)
    wb    = Workbook()

    DARK_F  = PatternFill("solid", fgColor="0F1421")
    AMBER_F = PatternFill("solid", fgColor="F59E0B")
    LIGHT_F = PatternFill("solid", fgColor="F8FAFC")
    WHITE_F = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="E2E8F0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, txt, bold=True, color="FFFFFF", fill=DARK_F):
        cell.value = txt
        cell.font  = Font(bold=bold, color=color, size=10)
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr

    def data_cell(cell, val, bold=False, fill=WHITE_F):
        cell.value  = val
        cell.font   = Font(bold=bold, size=10)
        cell.fill   = fill
        cell.border = bdr

    # ── Sheet 1: Survey Summary ──
    ws = wb.active
    ws.title = "Survey Summary"
    ws["A1"] = f"MineVisionAI — Volume Survey Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"{site.name} · {site.location or ''}, {site.state or ''}"
    ws["A2"].font = Font(size=11, color="64748B")
    ws["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}"
    ws["A3"].font = Font(size=10, color="94A3B8")
    for rng in ("A1:E1","A2:E2","A3:E3"):
        ws.merge_cells(rng)
    ws.append([])

    def add_section(ws, title, rows):
        ws.append([title])
        ws[ws.max_row][0].font = Font(bold=True, size=11)
        ws[ws.max_row][0].fill = PatternFill("solid", fgColor="F59E0B")
        for i, (k, v) in enumerate(rows):
            ws.append([k, v])
            r = ws.max_row
            fill = LIGHT_F if i % 2 == 0 else WHITE_F
            ws.cell(r,1).font  = Font(bold=True, size=10); ws.cell(r,1).fill = fill; ws.cell(r,1).border = bdr
            ws.cell(r,2).font  = Font(size=10);            ws.cell(r,2).fill = fill; ws.cell(r,2).border = bdr
        ws.append([])

    add_section(ws, "Site Information", [
        ("Site Name",     site.name),
        ("Location",      f"{site.location or NA}, {site.state or NA}"),
        ("Mine Type",     site.mine_type or NA),
        ("Coordinates",   f"{site.latitude or NA}°N, {site.longitude or NA}°E"),
        ("Area",          _v(site.area_km2, "%.4f", " km²")),
        ("Elevation Min", _v(site.elevation_min, "%.1f", " m ASL")),
        ("Elevation Max", _v(site.elevation_max, "%.1f", " m ASL")),
    ])

    add_section(ws, "Volume Results", [
        ("Stockpile Volume",    _v(survey.stockpile_volume, "%,.0f", " m³")),
        ("Cut Volume",          _v(survey.cut_volume,       "%,.0f", " m³")),
        ("Net Change",          _v(survey.net_change,       "%+,.0f", " m³") if survey.net_change else NA),
        ("Reference Plane",     _v(survey.reference_elev,   "%.2f",  " m ASL")),
        ("Volume Method",       "TIN above reference plane"),
    ])

    add_section(ws, "Processing Details", [
        ("Drone Model",         survey.drone_model or NA),
        ("Flying Height",       _v(survey.flying_height_m, "%.0f", " m")),
        ("Images Uploaded",     _v(survey.image_count, "%d")),
        ("Images Passed QC",    _v(survey.images_passed, "%d")),
        ("Images Rejected",     _v(survey.images_rejected, "%d")),
        ("GCPs Used",           _v(survey.gcp_count, "%d")),
        ("GCP RMSE",            _v(survey.gcp_rmse, "%.2f", " cm")),
        ("DEM Resolution",      _v(survey.dem_resolution, "%.2f", " m/px")),
        ("Coordinate System",   survey.dem_crs or NA),
    ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    # ── Sheet 2: Image QC ──
    ws2 = wb.create_sheet("Image QC")
    headers = ["Filename","Passed","Blur Score","Geotagged","Lat","Lon","Alt (m)","Reason"]
    for col, h in enumerate(headers, 1):
        hdr(ws2.cell(1, col), h)

    img_files = db.query(UploadedFile).filter(
        UploadedFile.survey_id == survey.id,
        UploadedFile.file_type == "image"
    ).all()
    for i, f in enumerate(img_files):
        fill = LIGHT_F if i % 2 == 0 else WHITE_F
        row = [
            f.filename,
            "Yes" if f.passed_qc else "No",
            f.blur_score,
            "Yes" if f.is_geotagged else "No",
            f.gps_lat, f.gps_lon, f.gps_alt,
            "OK" if f.passed_qc else "Failed QC",
        ]
        for col, val in enumerate(row, 1):
            data_cell(ws2.cell(i+2, col), val, fill=fill)
    for c in "ABCDEFGH":
        ws2.column_dimensions[c].width = 18

    # ── Sheet 3: Volume History ──
    ws3 = wb.create_sheet("Volume History")
    for col, h in enumerate(["Month","Volume (m³)","Change vs Previous"], 1):
        hdr(ws3.cell(1, col), h)

    hist = db.query(VolumeHistory).filter(
        VolumeHistory.site_id == site.id
    ).order_by(VolumeHistory.id).all()
    prev = None
    for i, v in enumerate(hist):
        fill = LIGHT_F if i % 2 == 0 else WHITE_F
        change = f"{v.volume_m3 - prev:+,.0f}" if prev is not None else "—"
        for col, val in enumerate([v.month_label, f"{v.volume_m3:,.0f}", change], 1):
            data_cell(ws3.cell(i+2, col), val, fill=fill)
        prev = v.volume_m3
    for c in "ABC":
        ws3.column_dimensions[c].width = 20

    # ── Sheet 4: Pipeline Log ──
    ws4 = wb.create_sheet("Pipeline Log")
    for col, h in enumerate(["Step","Status","Detail"], 1):
        hdr(ws4.cell(1, col), h)
    steps = json.loads(survey.pipeline_steps) if survey.pipeline_steps else []
    for i, s in enumerate(steps):
        fill = LIGHT_F if i % 2 == 0 else WHITE_F
        for col, val in enumerate([s["name"], s["status"].title(), s.get("detail","")], 1):
            data_cell(ws4.cell(i+2, col), val, fill=fill)
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 60

    wb.save(path)
    return path
